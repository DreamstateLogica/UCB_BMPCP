#!/usr/bin/env python3
"""
A script to selectively pull desired data from ORCA .out files (neutral,
anion, cation) within a series of subdirectories and compile the results into
a single Excel file with embedded statistical analysis formulas.
"""
__author__ = "Nabil Mouhajir, Peter Waddell"
__copyright__ = "Copyright 2024"
__credits__ = ["Nabil Mouhajir, Peter Waddell"]
__version__ = "1.0"
__date__ = "2025/07/22"
__maintainer__ = "Nabil Mouhajir"
__email__ = "mouhajirni@hotmail.com"
__status__ = "Prototype"

import os
import sys
import pandas as pd

# --- FIX FOR ModuleNotFoundError ---
try:
    src_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(src_dir))
    if project_root not in sys.path:
        sys.path.insert(0, project_root)
except NameError:
    sys.path.insert(0, os.path.abspath('../..'))
# --- END FIX ---

from orca_data_extraction.src.structure_data_builder import StructureDataBuilder
from orca_data_extraction.src.orca_out_to_json import make_json_list

def create_analysis_workbook(data_collections, output_basename):
    """
    Writes data collections to a multi-sheet .xlsx file, including a blank
    column for experimental yields and embedded formulas for R^2 and MAE.
    Yields are linked across sheets.

    Parameters
    ----------
    data_collections : dict
        A dictionary where keys are sheet names ('neutral', 'anion', 'cation')
        and values are lists of StructureData instances.
    output_basename : str
        Name of the output file, without extension.
    """
    if not any(data_collections.values()):
        print("No data was successfully extracted. Excel file will not be created.")
        return
    
    output_filename = output_basename + '.xlsx'

    try:
        writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    except ImportError:
        print("\nError: The 'openpyxl' library is required to write Excel files.")
        print("Please install it using: pip install openpyxl")
        return

    # Write each collection to a separate sheet
    for calc_type, sd_list in data_collections.items():
        if not sd_list:
            print(f"No data for '{calc_type}', skipping sheet.")
            continue

        df = pd.json_normalize(make_json_list(sd_list))
        df.insert(1, "Experimental Yield (%)", "")
        
        if 'origin' in df.columns:
            cols = ['origin'] + [col for col in df.columns if col != 'origin']
            df = df[cols]
        
        sheet_name = calc_type.capitalize()
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # --- Add Formulas and Links ---
    workbook = writer.book
    
    # Link the yield columns if multiple sheets exist
    sheet_names = [s.capitalize() for s in data_collections.keys() if data_collections[s]]
    if len(sheet_names) > 1:
        master_sheet = "Neutral" # Designate Neutral as the primary input sheet
        if master_sheet not in sheet_names: # Fallback if no neutral data
             master_sheet = sheet_names[0]

        for sheet_name in sheet_names:
            if sheet_name == master_sheet:
                continue
            
            worksheet = workbook[sheet_name]
            num_rows = len(data_collections[sheet_name.lower()])
            for i in range(2, num_rows + 2):
                worksheet[f'B{i}'].value = f"={master_sheet}!B{i}"
    
    # Add statistical formulas to each sheet
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        num_rows = len(data_collections[sheet_name.lower()])
        num_cols = worksheet.max_column
        
        stats_start_row = num_rows + 4
        worksheet.cell(row=stats_start_row - 1, column=1, value="Correlation Analysis:")
        worksheet.cell(row=stats_start_row, column=1, value="R^2 vs. Yield")
        worksheet.cell(row=stats_start_row + 1, column=1, value="MAE vs. Yield")

        yield_col = 'B'
        data_start_col_index = 3  # Column 'C'

        from openpyxl.utils import get_column_letter
        for i in range(data_start_col_index, num_cols + 1):
            col_letter = get_column_letter(i)
            
            yield_range = f'{yield_col}2:{yield_col}{num_rows + 1}'
            descriptor_range = f'{col_letter}2:{col_letter}{num_rows + 1}'

            r2_formula = f'=IF(COUNT({yield_range})<2, "Enter Yields", IFERROR(RSQ({yield_range}, {descriptor_range}), "NA"))'
            mae_formula = f'=IF(COUNT({yield_range})<1, "Enter Yields", IFERROR(SUMPRODUCT(ABS({yield_range}-{descriptor_range}))/COUNT({yield_range}), "NA"))'

            worksheet.cell(row=stats_start_row, column=i, value=r2_formula)
            worksheet.cell(row=stats_start_row + 1, column=i, value=mae_formula)

    writer.close()
    print(f'\nProcess complete! Results and analysis formulas saved as "{output_filename}"')


def main():
    valid_types = ['neutral', 'anion', 'cation', 'all']
    if len(sys.argv) != 4 or sys.argv[3].lower() not in valid_types:
        print("Usage: python orca_out_to_xlsx.py <input_json_file> <output_excel_name> <type>")
        print("  <type> must be one of: neutral, anion, cation, all")
        sys.exit(1)

    inputs_name = sys.argv[1]
    output_basename = sys.argv[2]
    calc_type_arg = sys.argv[3].lower()

    execution_path = os.getcwd()
    json_input_path = os.path.join(execution_path, inputs_name)

    if not os.path.isfile(json_input_path):
        print(f"Error: Input JSON file '{json_input_path}' not found.")
        quit()

    print(f"Using input file: {json_input_path}\n")
    
    structure_data_builder = StructureDataBuilder(json_input_path)
    
    types_to_process = valid_types[:-1] if calc_type_arg == 'all' else [calc_type_arg]
    data_collections = {t: [] for t in types_to_process}

    subdirs = sorted([d for d in os.listdir(execution_path) if os.path.isdir(d) and d.startswith('mol_')])
    if not subdirs:
        print("No subdirectories with prefix 'mol_' found in the current directory.")
        return

    for dirname in subdirs:
        dirpath = os.path.join(execution_path, dirname)
        print(f"--- Searching in directory: {dirpath} ---")
        
        for calc_type in types_to_process:
            found_out_file = False
            # Look for a file that contains the type keyword, e.g., 'neutral'
            for filename in os.listdir(dirpath):
                if calc_type in filename and filename.endswith(".out"):
                    out_filepath = os.path.join(dirpath, filename)
                    print(f"Found '{calc_type}' file: {filename}")
                    
                    try:
                        sd_object = structure_data_builder.build(out_filepath)
                        sd_object.origin = dirname
                        data_collections[calc_type].append(sd_object)
                    except Exception as e:
                        print(f"  -> An error occurred while processing {filename}: {e}")
                    
                    found_out_file = True
                    break
            
            if not found_out_file:
                print(f"Warning: No '{calc_type}' .out file found in {dirname}")

        print("-" * (len(dirpath) + 24))

    create_analysis_workbook(data_collections, output_basename)


if __name__ == '__main__':
    main()